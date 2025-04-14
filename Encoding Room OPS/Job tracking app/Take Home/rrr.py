import sys
from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QFormLayout,
    QLineEdit, QPushButton, QMessageBox, QCheckBox
)
from openpyxl import load_workbook

class ExcelForm(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Excel Form Filler')

        layout = QVBoxLayout()
        form_layout = QFormLayout()

        # Input fields matching your Excel form:
        self.customer = QLineEdit()
        self.part_num = QLineEdit()
        self.job_ticket_num = QLineEdit()
        self.customer_po = QLineEdit()
        self.inlay_type = QLineEdit()
        self.label_size = QLineEdit()
        self.layout_field = QLineEdit()
        self.qty = QLineEdit()

        form_layout.addRow("Customer:", self.customer)
        form_layout.addRow("Part #:", self.part_num)
        form_layout.addRow("Job Ticket #:", self.job_ticket_num)
        form_layout.addRow("Customer PO #:", self.customer_po)
        form_layout.addRow("Inlay Type:", self.inlay_type)
        form_layout.addRow("Label Size:", self.label_size)
        form_layout.addRow("Layout:", self.layout_field)
        form_layout.addRow("QTY:", self.qty)

        # Checkbox example
        self.verify_customer_cb = QCheckBox("Verify Customer name?")
        form_layout.addRow(self.verify_customer_cb)

        # Submit Button
        submit_btn = QPushButton("Generate Excel")
        submit_btn.clicked.connect(self.generate_excel)

        layout.addLayout(form_layout)
        layout.addWidget(submit_btn)
        self.setLayout(layout)

    def generate_excel(self):
        # Load workbook and select active worksheet
        wb = load_workbook(r"C:\Users\Encoding 3\Desktop\Encoding Checklist New WIPv2.xlsx")
        ws = wb.active

        try:
            # Populate Excel cells with form data (adjust cell references as your actual form layout)
            ws["C2"] = self.customer.text()
            ws["C3"] = self.part_num.text()
            ws["C4"] = self.job_ticket_num.text()
            ws["C5"] = self.customer_po.text()
            ws["C6"] = self.inlay_type.text()
            ws["C7"] = self.label_size.text()
            ws["C8"] = self.layout_field.text()
            ws["C9"] = self.qty.text()

            # Checkbox logic for cell A12:
            ws["A12"] = "✓" if self.verify_customer_cb.isChecked() else "☐"

            # Save filled form as new file
            filled_form_path = r"C:\Users\Encoding 3\Desktop\Encoding_Checklist_Filled.xlsx"
            wb.save(filled_form_path)

            QMessageBox.information(self, "Success", f"Form saved successfully!\n\n{filled_form_path}")

        except Exception as e:
            QMessageBox.warning(self, "Error", f"An error occurred:\n{str(e)}")

# Main execution
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelForm()
    window.show()
    sys.exit(app.exec())
